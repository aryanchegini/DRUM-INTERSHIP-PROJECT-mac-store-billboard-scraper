from mac_store_scraper import MacStoreScraper
import openpyxl
import pandas

#declare the list of postcodes here so that this script doesn't have to run 800 times
wb = openpyxl.load_workbook('M.A.C Holiday 21 - OOH V5 - BOOKED.xlsx')
ws = wb['ATLAS BOOKED SITE LIST']
list_of_postcodes = []
for cell in ws['H']:
    if cell.value != None:
        list_of_postcodes.append(cell.value)
list_of_postcodes.pop(0)



class BillBoardFinder:
    def __init__(self, postcode_list:list = list_of_postcodes) -> None:
        self.excelfile = 'M.A.C Holiday 21 - OOH V5 - BOOKED.xlsx'
        self.wb = openpyxl.load_workbook(self.excelfile)
        self.main_sheet = self.wb['ATLAS BOOKED SITE LIST']
        self.index = 0
        self.df = pandas.read_excel(self.excelfile)
        self.list_of_postcodes = postcode_list
    
    @property
    def verify_postcode(self) -> bool:
        if self.bb_postcode in self.list_of_postcodes:
            return True
        elif self.bb_postcode not in self.list_of_postcodes:
            return False

    def find_postcode_nearest(self) -> str:
        scraper = MacStoreScraper(self.list_of_postcodes[self.index])
        while not scraper.scraped:
            try:
                nearest_postcode = scraper.give_store_postcode()
                scraper.scraped = True
            except Exception:
                pass
        # nearest_postcode = scraper.give_store_postcode()
        return nearest_postcode

    def insert_postcodes(self) -> None:
        nearest_postcodes = []
        for i in range(len(self.list_of_postcodes)):
            self.index = i
            nearest_postcode = self.find_postcode_nearest()
            nearest_postcodes.append(nearest_postcode)
            #timing
            print(f'Index {self.index} ')
            print(nearest_postcode)
        self.df["Nearest M.A.C Store Postcode"] = nearest_postcodes
        self.df.to_excel(r'./M.A.C Holiday 21 - OOH V5 - BOOKED - With Nearest Store Postcodes.xlsx', sheet_name='ATLAS BOOKED SITE LIST')
        return
        
    
test1 = BillBoardFinder(postcode_list=list_of_postcodes)
test1.insert_postcodes()